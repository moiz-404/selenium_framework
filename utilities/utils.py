# import softest
import logging
import os
import inspect
import openpyxl
import csv
from openpyxl.styles import PatternFill
from pathlib import Path
import yaml
from yaml import SafeLoader


# class Utils(softest.TestCase):
class Utils():
    def assertListItemText(self, list, value):
        for i in list:
            print("The texts is :" + i.text)
            assert i.text == value
            print("assert pass")


    # def assertListItemTextWithSoftest(self, list, value):
    #     for i in list:
    #         print("The texts is :" + i.text)
    #         self.soft_assert(self.assertEqual, i.text, value)
    #         if i.text == value:
    #             print("assert passed")
    #         else:
    #             print("assert Failed")
    #
    #     self.assert_all()

#######################################  logger  ######################################################
    def logger(logLevel = logging.DEBUG):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        filehandler = logging.FileHandler(filename="./logs/automation.log")
        formatter = logging.Formatter('%(asctime)s,%(msecs)d %(name)s - %(levelname)s : %(message)s',
                                      datefmt='%m/%d/%y %H:%M:%S %p')
        filehandler.setFormatter(formatter)
        logger.addHandler(filehandler)
        return logger

#######################################  excel file  ######################################################

    def getRowCount(file,sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_row)



    def getColumnCount(file,sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_column)


    def readData(file,sheetName,rowNo,colsNo):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(rowNo, colsNo).value


    def readAllData(filename,sheetName):
        dataList =[]
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetName]
        row_ct = sheet.max_row
        cols_ct = sheet.max_column
        for r in range(2, row_ct+1):
            row = []
            for c in range(2, cols_ct + 1):
                row.append(sheet.cell(row_ct=r, cols_ct=c).value)
            dataList.append(row)
        return dataList


    def writeData(file,sheetName,rowNo,colsNo,data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(rowNo, colsNo).value = data
        workbook.save(file)


    def fillGreenColor(file,sheetName,rowNo,colsNo):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        greenFill = PatternFill(start_color='68b212',
                                end_color='60b212',
                                fill_type='solid')
        sheet.cell(rowNo, colsNo).value = greenFill
        workbook.save(file)


    def fillRedColor(file, sheetName, rowNo, colsNo):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        redFill = PatternFill(start_color='ff0000',
                                end_color='ff0000',
                                fill_type='solid')
        sheet.cell(rowNo, colsNo).value = redFill
        workbook.save(file)

#######################################  csv file  ######################################################

    def readAllDataFromcsv(DATA_FILE):
        # dataList = []
        # csvdata = open(filename, 'r')
        # reader =csv.reader(csvdata)
        # # next(reader)
        # for rows in reader:
        #     dataList.append(rows)
        # return dataList

        with open(DATA_FILE, 'r') as csvdata:
            reader = csv.reader(csvdata)
            next(reader)
            dataList = [tuple(rows) for rows in reader]
        return dataList

#######################################  yaml file  ######################################################

    def readAllDataFromYaml(DATA_FILE):
        with open(DATA_FILE, "r") as fh:
            data = yaml.load(fh, Loader=SafeLoader)
        return data

